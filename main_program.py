# python
# Created by Tuan Anh Phan on 18.04.2023

from os import getcwd as os_getcwd
from os import path as os_path
from os import chdir as os_chdir
from sys import argv as sys_argv
from sys import exit as sys_exit
from time import localtime
from random import shuffle
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from openpyxl import open as open_ex


from bytearray_text_kuz import blockDecrypt, func_RKey


"""
    All configure data in file config.txt
"""
size_Width = 1366
size_Height = 768
size_Push_button_W = 100
size_Push_button_H = 75

os_chdir(os_path.join(os_getcwd(), "src"))  # change directory to src
class Window(QWidget):
    def __init__(self):
        QWidget.__init__(self)
        # --------- INIT VARS -----------
        self.time_end = None
        self.fullPathExcel = None
        self.counter_wrong_answer = 2  # init // dem so dau hoi sai de index vao diag + 2 dong thong tin
        self.layout_scrollbar_res = None
        self.scroll_bar_res = None
        self.layout_diag_result = None
        self.diag_result = None
        self.layout_GUI_TEST = QGridLayout()
        self.frameRight = None
        self.rand_num = 0
        self.questions = None
        self.file_name_result = None
        self.counter_num_question = 1
        self.bottom_main_frame = None
        self.result_file = None
        self.button_next_ingame = None
        self.label_infor = None
        self.frameLeft = None
        self.layout_frame_left = None
        self.layout_frame_right = None
        self.all_Answer_button = None
        self.correct_answer = None
        self.Dial_Game = None
        self.layout_menu_home = None
        self.button_reset = None
        self.group = None
        self.file_name = None
        self.name = None
        self.button_start = None
        self.button_exit = None
        self.time_st = None
        self.question = None
        self.answer_1 = None
        self.answer_2 = None
        self.answer_3 = None
        self.answer_4 = None
        self.score = 0
        # --------- END INIT VARS -----------

        # --------- CONFIG AREA ---------
        # ---- FILE CONFIG
        config_check_file_path = os_path.join(os_getcwd(), "config_path_file.txt")
        data_file_check = open(config_check_file_path, "r", encoding='utf-8').read()
        if data_file_check == "":
            config_file_path = os_path.join(os_getcwd(), "config.txt")
        else:
            config_file_path = data_file_check
        self.checkFilePath(config_file_path)
        # arr_data_config = open(config_file_path, "r", encoding='utf-8').read().split("\n")
        # config_file_path = os_path.join(os_getcwd(), "config.txt")
        # ---- END FILE CONFIG

        self.arr_configData = open(config_file_path, "r", encoding='utf-8').read().split("\n")
        self.num_question = int(self.arr_configData[5].split("=")[1])
        self.font = QFont(self.arr_configData[0].split("=")[1], int(self.arr_configData[1].split("=")[1]))
        self.setFont(self.font)

        path_file = os_path.join("data", self.arr_configData[2].split("=")[1])
        lst_data_from_file = open(path_file, "r", encoding='utf-8').read().split("\n")
        while "" in lst_data_from_file:
            lst_data_from_file.remove("")
        if lst_data_from_file[0] != "ENCRYPTED_":
            messageBox = QMessageBox()
            messageBox.setWindowTitle("Ошибка")
            messageBox.setIcon(QMessageBox.Critical)
            messageBox.setText("Выбрали незашифронное слово")
            messageBox.setStandardButtons(QMessageBox.Close)
            messageBox.buttonClicked.connect(messageBox.close)
            retval = messageBox.exec_()
            exit(0)
        else:
            lst_data_from_file.pop(0)
        key = "СпецФак202212345678764321"
        self.rKey = func_RKey(bytearray(key.encode()))

        self.questions_answers = self.decryptData(lst_data_from_file)
        # delete "" in dict
        # while "" in self.questions_answers:
        #     self.questions_answers.pop("")
        # while "\n" in self.questions_answers:
        #     self.questions_answers.pop("\n")

        self.questions = list(self.questions_answers.keys())
        shuffle(self.questions)
        # ------- END CONFIG AREA ---------

        # -------- INIT DIAG LAYOUT -------
        self.diag_result = QDialog()
        self.diag_result.setMinimumSize(size_Width, size_Height)
        self.diag_result.setFont(self.font)
        self.diag_result.setMinimumSize(700, 700)

        self.layout_diag_result = QGridLayout()
        self.diag_result.setLayout(self.layout_diag_result)
        # -------- END INIT DIAG LAYOUT -------

        self.setWindowTitle("------ПРОГРАММА ТЕСТИРОВАНИЯ ПО РУССКОМУ ЯЗЫКУ------")
        self.resize(size_Width, size_Height)
        self.setMinimumWidth(size_Width)
        self.setMinimumHeight(size_Height)
        self.menu_home()

    def menu_home(self):
        self.layout_menu_home = QGridLayout()
        self.setLayout(self.layout_menu_home)

        self.name = QLineEdit()
        self.name.setMinimumSize(100, 100)
        self.name.setPlaceholderText("Введите Ф.И.О.")

        self.name.setFont(self.font)

        self.group = QLineEdit()
        self.group.setPlaceholderText("Введите номер группы")
        self.group.setMinimumSize(100, 100)
        self.group.setFont(self.font)

        self.bottom_main_frame = QFrame()
        self.layout_menu_home.addWidget(self.bottom_main_frame, 3, 1)
        layout_main_frame = QHBoxLayout()
        self.bottom_main_frame.setLayout(layout_main_frame)

        self.button_start = QPushButton("Начать")
        self.button_start.clicked.connect(self.check_info_to_start)

        self.button_exit = QPushButton("Выйти")
        self.button_exit.clicked.connect(QApplication.instance().quit)

        self.layout_menu_home.addWidget(self.name, 1, 1)
        self.layout_menu_home.addWidget(QLabel("Ф.И.О.: "), 1, 0)
        self.layout_menu_home.addWidget(self.group, 2, 1)
        self.layout_menu_home.addWidget(QLabel("Группа: "), 2, 0)
        layout_main_frame.addWidget(self.button_start)
        layout_main_frame.addWidget(self.button_exit)

    def test_GUI(self):
        self.Dial_Game = QDialog()
        self.Dial_Game.setFont(self.font)
        self.Dial_Game.setMinimumSize(size_Width, size_Height)
        self.Dial_Game.setLayout(self.layout_GUI_TEST)
        self.config_frameLeft()
        self.config_frameRight()

        self.label_infor = QLabel(
            "Слушатель: {}\nГруппа: {}\nБалл: {}".format(self.name.text(), self.group.text(), self.score))
        self.label_infor.setStyleSheet("border: 1px solid black;")
        self.label_infor.setFont(self.font)
        self.layout_GUI_TEST.addWidget(self.label_infor, 1, 0)

        self.button_next_ingame = QPushButton("Следующий вопрос")
        self.button_next_ingame.setEnabled(False)
        self.button_next_ingame.setFont(self.font)
        self.button_next_ingame.clicked.connect(self.on_clicked_next_ques)
        self.layout_GUI_TEST.addWidget(self.button_next_ingame, 1, 1)

        self.temp_gen_ques()
        retval = self.Dial_Game.exec_()

    def on_clicked_next_ques(self):
        self.full_clearLayout()
        self.temp_gen_ques()

    def config_frameLeft(self):
        self.frameLeft = QFrame()
        self.layout_GUI_TEST.addWidget(self.frameLeft, 0, 0)
        self.layout_frame_left = QVBoxLayout()
        self.frameLeft.setLayout(self.layout_frame_left)

    def config_frameRight(self):
        self.frameRight = QFrame()
        self.frameRight.setMaximumSize(530, 500)
        self.layout_GUI_TEST.addWidget(self.frameRight, 0, 1)
        self.layout_frame_right = QVBoxLayout()
        self.frameRight.setLayout(self.layout_frame_right)

    def temp_gen_ques(self):
        self.button_next_ingame.setEnabled(False)
        self.all_Answer_button = QButtonGroup()
        self.all_Answer_button.setExclusive(False)
        self.all_Answer_button.buttonClicked[int].connect(self.on_button_clicked)

        self.question = QLabel()
        self.layout_frame_left.addWidget(self.question)

        self.answer_1 = QPushButton()
        self.answer_1.setMinimumSize(size_Push_button_W, size_Push_button_H)
        self.answer_2 = QPushButton()
        self.answer_2.setMinimumSize(size_Push_button_W, size_Push_button_H)
        self.answer_3 = QPushButton()
        self.answer_3.setMinimumSize(size_Push_button_W, size_Push_button_H)
        self.answer_4 = QPushButton()
        self.answer_4.setMinimumSize(size_Push_button_W, size_Push_button_H)

        self.all_Answer_button.addButton(self.answer_1)
        self.all_Answer_button.addButton(self.answer_2)
        self.all_Answer_button.addButton(self.answer_3)
        self.all_Answer_button.addButton(self.answer_4)

        self.layout_frame_right.addWidget(self.answer_1)
        self.layout_frame_right.addWidget(self.answer_2)
        self.layout_frame_right.addWidget(self.answer_3)
        self.layout_frame_right.addWidget(self.answer_4)

        # ------ GEN QUESTIONS ---------------
        self.rand_num += 1
        self.correct_answer = self.questions_answers[self.questions[self.rand_num]]
        wrong_answer = list(self.questions_answers.values())
        del wrong_answer[wrong_answer.index(self.correct_answer)]
        answer_Option_full = [self.correct_answer] + wrong_answer
        answer_Option = answer_Option_full[:4]
        shuffle(answer_Option)
        temp_Pixmap_obj = QPixmap(
            os_path.join(os_getcwd(), "image", self.arr_configData[3].split("=")[1],
                         str(self.questions[self.rand_num]).replace('\n', '')))
        Pixmap_obj = temp_Pixmap_obj.scaled(750, 750)
        self.question.setPixmap(Pixmap_obj)
        self.answer_1.setText(str(answer_Option[0]))
        self.answer_2.setText(str(answer_Option[1]))
        self.answer_3.setText(str(answer_Option[2]))
        self.answer_4.setText(str(answer_Option[3]))

        # ------ END GEN QUESTIONS ---------------

    def check_info_to_start(self):
        temp_time = localtime()
        date = "{}.{}.{}".format(temp_time.tm_mday, temp_time.tm_mon, temp_time.tm_year)
        self.time_st = "{}ч_{}м_".format(temp_time.tm_hour, temp_time.tm_min) + date

        if self.name.text() != "" and self.group.text() != "":
            self.name.setReadOnly(True)
            self.group.setReadOnly(True)
            self.test_GUI()
        else:
            self.reset_all_lineEdit()

    def reset_all_lineEdit(self):
        self.name.clear()
        self.name.setText("")
        self.name.setReadOnly(False)

        self.group.clear()
        self.group.setText("")
        self.group.setReadOnly(False)

        self.button_start.setEnabled(True)
        self.counter_num_question = 1
        self.score = 0

    def func_diag_result(self):
        label_name = QLabel("Слушатель: %s" % self.name.text())
        label_name.setStyleSheet("border: 1px solid black;")
        label_group = QLabel("Группа: %s" % self.group.text())
        label_group.setStyleSheet("border: 1px solid black;")
        label_score = QLabel("Балл: %s" % self.score)
        label_score.setStyleSheet("border: 1px solid black;")

        self.layout_diag_result.addWidget(label_name, 0, 0)
        self.layout_diag_result.addWidget(label_group, 0, 1)
        self.layout_diag_result.addWidget(label_score, 0, 2)

        label_res_pic = QLabel("Рисунок")
        label_res_pic.setStyleSheet("font-weight: bold;"
                                    "font-size=20;")
        label_res_cor_ans = QLabel("Правильные ответы")
        label_res_cor_ans.setStyleSheet("font-weight: bold;"
                                        "font-size=20;")
        label_res_ans = QLabel("Ваши ответы")
        label_res_ans.setStyleSheet("font-weight: bold;"
                                    "font-size=20;")

        self.layout_diag_result.addWidget(label_res_pic, 1, 0)
        self.layout_diag_result.addWidget(label_res_cor_ans, 1, 1)
        self.layout_diag_result.addWidget(label_res_ans, 1, 2)

        close_button = QPushButton("Закрыть")
        self.layout_diag_result.addWidget(close_button, self.counter_wrong_answer + 1, 1)
        close_button.clicked.connect(self.diag_result.close)

        retval = self.diag_result.exec_()

    def on_button_clicked(self, id_button):
        self.counter_num_question += 1

        self.answer_1.setEnabled(False)
        self.answer_2.setEnabled(False)
        self.answer_3.setEnabled(False)
        self.answer_4.setEnabled(False)
        self.button_next_ingame.setEnabled(True)

        for button in self.all_Answer_button.buttons():
            if button is self.all_Answer_button.button(id_button):
                if self.correct_answer == button.text():
                    self.score += 1
                    button.setStyleSheet("color : green")
                else:
                    button.setStyleSheet("color : red")

                    # process image and data
                    self.score = self.score
                    pixmap_temp = self.question.pixmap().scaled(100, 100)
                    label_temp = QLabel()
                    label_temp.setPixmap(pixmap_temp)

                    label_temp_correctAns = QLabel(self.correct_answer)
                    label_temp_correctAns.setFont(self.font)
                    label_temp_correctAns.setStyleSheet("color : green")

                    label_temp_WrongAns = QLabel(button.text())
                    label_temp_WrongAns.setFont(self.font)
                    label_temp_WrongAns.setStyleSheet("color : red")
                    # end process image and data

                    self.layout_diag_result.addWidget(label_temp, self.counter_wrong_answer, 0)
                    self.layout_diag_result.addWidget(label_temp_correctAns, self.counter_wrong_answer, 1)
                    self.layout_diag_result.addWidget(label_temp_WrongAns, self.counter_wrong_answer, 2)
                    self.counter_wrong_answer += 1

        self.label_infor.setText(
            "Слушатель: {}\nГруппа: {}\nБалл: {}".format(self.name.text(), self.group.text(), self.score))

        if self.counter_num_question > self.num_question:
            self.on_clicked_close_diag()

    def full_clearLayout(self):
        while self.layout_frame_left.count():
            child = self.layout_frame_left.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        while self.layout_frame_right.count():
            child = self.layout_frame_right.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

    def on_clicked_close_diag(self):
        self.Dial_Game.close()
        self.func_diag_result()
        self.func_writeExcel()
        self.button_start.setEnabled(False)

    def func_writeExcel(self):
        pathExcel = "res.xlsx"
        fullPathExcel = os_path.join(os_getcwd(), "result", pathExcel)
        wb = open_ex(fullPathExcel)
        sheet = wb["Sheet1"]

        sheet["A1"].value = "Ф. И. О. слушателей"
        sheet["B1"].value = "Группа"
        sheet["C1"].value = "Балл"
        sheet["D1"].value = "Время начала"
        sheet["E1"].value = "Время окончания"

        max_row = sheet.max_row
        sheet.cell(row=max_row + 1, column=1, value=str(self.name.text()))
        sheet.cell(row=max_row + 1, column=2, value=str(self.group.text()))
        sheet.cell(row=max_row + 1, column=3, value=int(self.score))
        sheet.cell(row=max_row + 1, column=4, value=str(self.time_st))

        temp_time = localtime()
        data_date = "{}.{}.{}".format(temp_time.tm_mday, temp_time.tm_mon, temp_time.tm_year)
        self.time_end = "{}ч_{}м_".format(temp_time.tm_hour, temp_time.tm_min) + data_date
        sheet.cell(row=max_row + 1, column=5, value=str(self.time_end))

        wb.save(fullPathExcel)
        wb.close()

    @staticmethod
    def checkFilePath(file_patch):
        if not os_path.isabs(file_patch):
            print(file_patch + "not found")
            exit(0)

    def decryptData(self, arr_data_in: list) -> dict:
        lst_de_data = []
        for data in arr_data_in:
            lst_de_data.append(blockDecrypt(self.rKey, bytearray.fromhex(data)).decode())

        dict_data_out = {}
        assert len(arr_data_in) % 2 == 0
        for i in range(len(lst_de_data) // 2):
            dict_data_out[lst_de_data[(2 * i) + 1]] = lst_de_data[(2 * i)]
        return dict_data_out

    # @staticmethod
    # def get_maximum_rows(*, sheet_object):
    #     rows = 0
    #     for max_row, row in enumerate(sheet_object, 1):
    #         if not all(col.value is None for col in row):
    #             rows += 1
    #     return rows


if __name__ == '__main__':
    app = QApplication(sys_argv)
    screen = Window()
    screen.show()
    sys_exit(app.exec_())

